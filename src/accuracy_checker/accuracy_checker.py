import openpyxl


class AccuracyCalculator:
	"""
	Класс, обеспечивающий вычисление основных метрик качества детектирования объектов:
	TPR (True Positive Rate), FDR (False Detection Rate),
	Average Precision (AP) и Mean Average Precision (mAP) по нескольким классам.

	Пример использования:
		>>> calc = AccuracyCalculator(iou_threshold=0.5)
		>>> calc.load_groundtruths("groundtruths.xlsx")
		>>> calc.load_detections("detections.xlsx")
		>>> print(f"TPR: {calc.calculate_tpr():.4f}")
		>>> print(f"FDR: {calc.calculate_fdr():.4f}")
		>>> class_name = "CAR"
		>>> print(f"AP for {class_name}: {calc.calculate_average_precision(class_name):.4f}")
		>>> print(f"mAP: {calc.calculate_mean_average_precision():.4f}")
	"""
	def __init__(self, iou_threshold=0.5):
		"""
		Инициализация класса для вычисления средней точности (AP).

		:param iou_threshold: Порог Intersection over Union (IoU), используется для определения,
							  соответствует ли найденный объект истинной разметке.
		"""
		self.iou_threshold = iou_threshold  # Порог IoU
		self.groundtruths = {}  # Dict: Истинная разметка (группируется по классам)
		self.detections = {}  # Dict: Предсказания детектора (группируются по классам)

	def load_groundtruths(self, file_path):
		"""
		Загрузка истинной разметки из файла (.xlsx).

		:param file_path: Путь к файлу с истинной разметкой.
		"""
		self.groundtruths = self._parse_annotations(file_path)

	def load_detections(self, file_path):
		"""
		Загрузка предсказаний детектора из файла (.xlsx).

		:param file_path: Путь к файлу с предсказаниями.
		"""
		self.detections = self._parse_detections(file_path)

	def calculate_tpr(self):
		all_classes = self.groundtruths.keys()
		tp, fn = 0, 0
		for class_name in all_classes:
			detections = self.detections[class_name]
			groundtruths = self.groundtruths[class_name]

			# 1. Сортируем предсказания по достоверности
			all_detections = self._sort_detections_by_confidence(detections)

			# 2. Поиск соответствий между предсказаниями и истинной разметкой
			fn += sum(len(groundtruths.get(frame, [])) for frame in groundtruths)
			for frame_id, dets in all_detections.items():
				gts = groundtruths.get(frame_id, [])  # Список всех прямоугольников для кадра
				tp_det, _, fn_det = self._match_detections_to_groundtruth(dets, gts)
				tp += tp_det
				fn -= tp_det

		return tp / (tp + fn) if (tp + fn) else 0

	def calculate_fdr(self):
		all_classes = self.groundtruths.keys()
		tp, fp = 0, 0
		for class_name in all_classes:
			detections = self.detections[class_name]
			groundtruths = self.groundtruths[class_name]

			# 1. Сортируем предсказания по достоверности
			all_detections = self._sort_detections_by_confidence(detections)

			# 2. Поиск соответствий между предсказаниями и истинной разметкой
			for frame_id, dets in all_detections.items():
				gts = groundtruths.get(frame_id, [])  # Список всех прямоугольников для кадра
				tp_det, fp_det, _ = self._match_detections_to_groundtruth(dets, gts)
				tp += tp_det
				fp += fp_det

		return fp / (tp + fp) if (tp + fp) else 0

	def calculate_average_precision(self, class_name):
		"""
		Вычисляет среднюю точность (Average Precision, AP).

		Схема вычисления:
		1. Обнаруженные окаймляющие прямоугольники сортируются
		   в порядке убывания достоверности наличия в них объектов.
		2. Для каждого обнаруженного прямоугольника выполняется
		   поиск соответствия из разметки согласно условию IoU ≥ τ.
		3. Выполняется вычисление точности (Precision) и отклика (Recall).
		(4). Строится зависимость точности от отклика.
		5. Вычисляется площадь под графиком построенной зависимости (AP - Average Precision).

		Предположительно разметка детектора имеет следующий вид:
		0 CAR 232 128 290 168 0.77

		:return: Значение средней точности AP.
		"""
		if class_name not in self.detections or class_name not in self.groundtruths:
			return 0.0

		# Получаем предсказания и истинную разметку для заданного класса
		detections = self.detections[class_name]
		groundtruths = self.groundtruths[class_name]

		# 1. Сортируем предсказания по достоверности
		all_detections = self._sort_detections_by_confidence(detections)

		# 2. Поиск соответствий между предсказаниями и истинной разметкой
		tp, fp, fn = 0, 0, sum(len(groundtruths.get(frame, [])) for frame in groundtruths)
		all_tp, all_fp, all_fn = [], [], []
		for frame_id, dets in all_detections.items():
			gts = groundtruths.get(frame_id, [])  # Список всех прямоугольников для кадра
			tp_det, fp_det, fn_det = self._match_detections_to_groundtruth(dets, gts)
			tp += tp_det
			fp += fp_det
			fn -= tp_det
			all_tp.append(tp)
			all_fp.append(fp)
			all_fn.append(fn)

		# 3. Вычисляем precision и recall для всех точек
		precisions, recalls = self._calculate_precision_recall(all_tp, all_fp, all_fn, groundtruths, detections)

		# 4. Вычисляем площадь под графиком зависимости точности от отклика
		return self._compute_ap(precisions, recalls)

	def calculate_mean_average_precision(self):
		"""
		Вычисляет среднюю точность по всем классам (Mean Average Precision, mAP).

		:return: Значение средней точности (mAP) по всем классам.
		"""
		all_classes = self.groundtruths.keys()
		total_ap = 0
		for class_name in all_classes:
			total_ap += self.calculate_average_precision(class_name)
		return total_ap / len(all_classes) if all_classes else 0

	# ======= Приватные методы =======
	@staticmethod
	def _parse_annotations(file_path):
		"""
		Парсинг xlsx-файла с истинной разметкой.

		:param file_path: Путь к файлу с разметкой.
		:return: Словарь {class_name: {frame_id: [список ограничивающих прямоугольников]}}.
		"""
		annotations = {}
		try:
			workbook = openpyxl.load_workbook(file_path)
			sheet = workbook.active
			for row in sheet.iter_rows(values_only=True):
				frame_id, class_name, x1, y1, x2, y2 = row
				frame_id = int(frame_id)
				bbox = [float(x1), float(y1), float(x2), float(y2)]
				if class_name not in annotations:
					annotations[class_name] = {}
				if frame_id not in annotations[class_name]:
					annotations[class_name][frame_id] = []
				annotations[class_name][frame_id].append(bbox)
		except FileNotFoundError:
			print(f"Файл {file_path} не найден.")
		except Exception as e:
			print(f"Ошибка при чтении файла {file_path}: {e}")
		return annotations

	@staticmethod
	def _parse_detections(file_path):
		"""
		Парсинг xlsx-файла с предсказаниями детектора.

		:param file_path: Путь к файлу с предсказаниями.
		:return: Словарь {class_name: {frame_id: [список предсказаний]}}, где каждое предсказание содержит:
				 [x1, y1, x2, y2, confidence].
		"""
		detections = {}
		try:
			workbook = openpyxl.load_workbook(file_path)
			sheet = workbook.active
			for row in sheet.iter_rows(values_only=True):
				frame_id, class_name, x1, y1, x2, y2, confidence = row
				frame_id = int(frame_id)
				confidence = float(confidence)
				bbox = [float(x1), float(y1), float(x2), float(y2), confidence]
				if class_name not in detections:
					detections[class_name] = {}
				if frame_id not in detections[class_name]:
					detections[class_name][frame_id] = []
				detections[class_name][frame_id].append(bbox)
		except FileNotFoundError:
			print(f"Файл {file_path} не найден.")
		except Exception as e:
			print(f"Ошибка при чтении файла {file_path}: {e}")
		return detections

	@staticmethod
	def _calculate_iou(bbox1, bbox2):
		"""
		Вычисляет Intersection over Union (IoU) для двух прямоугольников.

		:param bbox1: Первый ограничивающий прямоугольник [x1, y1, x2, y2].
		:param bbox2: Второй ограничивающий прямоугольник [x1, y1, x2, y2].
		:return: Значение IoU (от 0 до 1).
		"""
		x1, y1, x2, y2 = bbox1
		x1g, y1g, x2g, y2g = bbox2

		xi1 = max(x1, x1g)
		yi1 = max(y1, y1g)
		xi2 = min(x2, x2g)
		yi2 = min(y2, y2g)

		inter_area = max(0, xi2 - xi1 + 1) * max(0, yi2 - yi1 + 1)
		bbox1_area = (x2 - x1 + 1) * (y2 - y1 + 1)
		bbox2_area = (x2g - x1g + 1) * (y2g - y1g + 1)
		union_area = bbox1_area + bbox2_area - inter_area

		return inter_area / union_area if union_area > 0 else 0

	def _match_detections_to_groundtruth(self, detections, groundtruths):
		"""
		Сопоставляет предсказания с истинной разметкой.

		:param detections: Список предсказаний для кадра.
		:param groundtruths: Список истинных объектов для кадра.
		:return: Количество TP (true positives), FP (false positives) и FN (false negatives).
		"""
		matched = set()
		tp, fp, fn = 0, 0, 0

		for det in detections:
			x1, y1, x2, y2, conf = det
			best_iou = 0
			best_gt_idx = -1
			# среди всех прямоугольников истинной разметки ищем тот, с которым наибольшее значение iou
			for idx, gt in enumerate(groundtruths):
				iou = self._calculate_iou([x1, y1, x2, y2], gt)
				if iou > best_iou:
					best_iou = iou
					best_gt_idx = idx

			if best_iou >= self.iou_threshold and best_gt_idx not in matched:
				tp += 1
				matched.add(best_gt_idx)
			else:
				# также сюда добавляются повторные детекции
				fp += 1

		fn = len(groundtruths) - len(matched)

		return tp, fp, fn

	@staticmethod
	def _sort_detections_by_confidence(detections):
		"""
		Сортирует предсказания по достоверности (confidence).

		:param detections: Словарь {frame_id: [список предсказаний]}.
		:return: Словарь {frame_id: [отсортированный список предсказаний]}.
		"""
		sorted_detections = {}
		for frame, dets in detections.items():
			sorted_detections[frame] = sorted(dets, key=lambda x: -x[-1])  # Сортируем по confidence (последний элемент)
		return sorted_detections

	@staticmethod
	def _calculate_precision_recall(all_tp, all_fp, all_fn, groundtruths, detections):
		"""
		Вычисляет массивы precision (точности) и recall (отклика).

		:param all_tp: Список TP (true positives).
		:param all_fp: Список FP (false positives).
		:return: Списки значений precision и recall.
		"""
		count_gt = [len(groundtruths.get(frame, [])) for frame in groundtruths]
		count_det = [len(groundtruths.get(frame, [])) for frame in detections]

		precisions = []
		recalls = []
		for tp, fp, fn, gt, det in zip(all_tp, all_fp, all_fn, count_gt, count_det):
			if gt > 0 and det > 0:
				precision = tp / (tp + fp)
				recall = tp / (tp + fn)
			elif gt == 0 and det > 0:  # tp == 0 nad fp > 0 and fn == 0
				precision = 0
				recall = 1
			elif gt > 0 and det == 0:  # tp == 0 nad fp == 0 and fn > 0
				precision = 1
				recall = 0
			elif gt == 0 and det == 0:  # tp == 0 nad fp == 0 and fn == 0
				precision = 1
				recall = 1

			precisions.append(precision)
			recalls.append(recall)
		return precisions, recalls

	@staticmethod
	def _compute_ap(precisions, recalls):
		"""
		Вычисляет площадь под кривой зависимости precision от recall.

		:param precisions: Список значений precision.
		:param recalls: Список значений recall.
		:return: Значение средней точности AP.
		"""
		precisions = [1.0] + precisions + [-1]
		recalls = [0.0] + recalls + [recalls[-1]]

		last_prec = 1
		last_rec = 0
		ap = 0
		for i in range(len(precisions)):
			if last_prec <= precisions[i]:
				continue
			else:
				ap += (last_prec * (recalls[i] - last_rec))
				last_rec = recalls[i]
				last_prec = precisions[i]

		return ap